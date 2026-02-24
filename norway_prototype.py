"""
norway_prototype.py  (version 6 – SQLite + Excel)
===================================================
Henter norske boligprisdata fra SSB og lagrer i:
  1. boligpriser.db  (SQLite-database)
  2. norsk_boligpris_prototype.xlsx (Excel)

Kjør:
    pip install requests openpyxl
    python norway_prototype.py
"""

import requests
import openpyxl
import sqlite3
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from itertools import product
from datetime import date

OUTPUT_EXCEL = "norsk_boligpris_prototype.xlsx"
OUTPUT_DB    = "boligpriser.db"
URL_REGION   = "https://data.ssb.no/api/v0/en/table/06035"
URL_KVARTAL  = "https://data.ssb.no/api/v0/en/table/07241"
ONSKEDE      = ["Oslo", "Bergen", "Trondheim", "Stavanger", "Hele landet", "The whole country"]

BLUE_DARK  = "1F4E79"
GREY_LIGHT = "F2F2F2"
WHITE      = "FFFFFF"

# ── HJELPEFUNKSJONER ──────────────────────────────────────────────────────────
def fix(t):
    return t.replace("km²","m²").replace("km2","m²").replace("KM2","m²")

def hent_meta(url):
    r = requests.get(url, timeout=30); r.raise_for_status()
    return {v["code"]: {"values": v["values"], "labels": v["valueTexts"]}
            for v in r.json()["variables"]}

def hent_json(url, query):
    r = requests.post(url, json={"query": query, "response": {"format": "json-stat2"}}, timeout=30)
    r.raise_for_status(); return r.json()

def parse(data):
    dims  = data["dimension"]
    vals  = data["value"]
    keys  = list(dims.keys())
    items = [list(dims[k]["category"]["label"].items()) for k in keys]
    rows, idx = [], 0
    for combo in product(*items):
        v = vals[idx] if idx < len(vals) else None
        idx += 1
        if v is not None:
            row = {keys[i]: combo[i][1] for i in range(len(keys))}
            row["verdi"] = v
            rows.append(row)
    return rows

# ── HENT DATA ─────────────────────────────────────────────────────────────────
def hent_region():
    print("Henter regiondata (årsvis, tabell 06035)...")
    meta = hent_meta(URL_REGION)
    reg_k, reg_l = [], []
    for k, l in zip(meta["Region"]["values"], meta["Region"]["labels"]):
        if any(l.lower().startswith(o.lower()) or o.lower() in l.lower() for o in ONSKEDE):
            reg_k.append(k); reg_l.append(l)
    if not reg_k:
        reg_k = meta["Region"]["values"][:5]; reg_l = meta["Region"]["labels"][:5]
    print(f"  Regioner: {reg_l}")
    innhold_navn = fix(meta["ContentsCode"]["labels"][0])
    query = [
        {"code": "Region",       "selection": {"filter": "item", "values": reg_k}},
        {"code": "Boligtype",    "selection": {"filter": "item", "values": meta["Boligtype"]["values"][:3]}},
        {"code": "ContentsCode", "selection": {"filter": "item", "values": [meta["ContentsCode"]["values"][0]]}},
        {"code": "Tid",          "selection": {"filter": "item", "values": meta["Tid"]["values"][-8:]}},
    ]
    rows = parse(hent_json(URL_REGION, query))
    print(f"  {len(rows)} rader")
    return rows, innhold_navn

def hent_kvartal():
    print("\nHenter kvartalsdata (nasjonalt, tabell 07241)...")
    meta = hent_meta(URL_KVARTAL)
    innhold_navn = fix(meta["ContentsCode"]["labels"][0])
    query = [
        {"code": "Boligtype",    "selection": {"filter": "item", "values": meta["Boligtype"]["values"][:3]}},
        {"code": "ContentsCode", "selection": {"filter": "item", "values": [meta["ContentsCode"]["values"][0]]}},
        {"code": "Tid",          "selection": {"filter": "item", "values": meta["Tid"]["values"][-12:]}},
    ]
    rows = parse(hent_json(URL_KVARTAL, query))
    print(f"  {len(rows)} rader")
    return rows, innhold_navn

# ── SQLITE ────────────────────────────────────────────────────────────────────
def spara_db(reg_rows, kv_rows):
    print(f"\nSparar til database ({OUTPUT_DB})...")
    conn = sqlite3.connect(OUTPUT_DB)
    c = conn.cursor()

    # Tabell 1: regiondata
    c.execute("""
        CREATE TABLE IF NOT EXISTS region_arsvis (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            region    TEXT,
            boligtype TEXT,
            periode   TEXT,
            pris_m2   REAL,
            land      TEXT DEFAULT 'Norge',
            kilde     TEXT DEFAULT 'SSB 06035',
            hentet    TEXT
        )
    """)

    # Tabell 2: kvartalsdata
    c.execute("""
        CREATE TABLE IF NOT EXISTS nasjonalt_kvartal (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            boligtype TEXT,
            periode   TEXT,
            pris_m2   REAL,
            land      TEXT DEFAULT 'Norge',
            kilde     TEXT DEFAULT 'SSB 07241',
            hentet    TEXT
        )
    """)

    hentet = date.today().isoformat()

    for row in reg_rows:
        c.execute("""
            INSERT INTO region_arsvis (region, boligtype, periode, pris_m2, hentet)
            VALUES (?, ?, ?, ?, ?)
        """, (row.get("Region"), row.get("Boligtype"), row.get("Tid"), row["verdi"], hentet))

    for row in kv_rows:
        c.execute("""
            INSERT INTO nasjonalt_kvartal (boligtype, periode, pris_m2, hentet)
            VALUES (?, ?, ?, ?)
        """, (row.get("Boligtype"), row.get("Tid"), row["verdi"], hentet))

    conn.commit()
    conn.close()

    # Bekreft innhold
    conn = sqlite3.connect(OUTPUT_DB)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM region_arsvis"); print(f"  region_arsvis: {c.fetchone()[0]} rader")
    c.execute("SELECT COUNT(*) FROM nasjonalt_kvartal"); print(f"  nasjonalt_kvartal: {c.fetchone()[0]} rader")
    conn.close()

# ── EXCEL ─────────────────────────────────────────────────────────────────────
def s(bold=True, farge=WHITE, size=10):
    return Font(name="Arial", bold=bold, color=farge, size=size)
def bg(c):
    return PatternFill("solid", start_color=c, fgColor=c)
def brd():
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)

def lag_ark(wb, tittel, rows, dim_keys, innhold_navn, tabell_navn, kilde):
    ws = wb.create_sheet(tittel)
    headers = dim_keys + [innhold_navn, "Land", "Kilde"]
    ws.row_dimensions[1].height = 28
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = s(); c.fill = bg(BLUE_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = brd()
    for row_idx, row in enumerate(rows, 2):
        rf = bg(GREY_LIGHT) if row_idx % 2 == 0 else bg(WHITE)
        vals = [row.get(k, "") for k in dim_keys] + [row["verdi"], "Norge", kilde]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.fill = rf; c.border = brd()
            c.font = Font(name="Arial", size=10)
            if col == len(dim_keys) + 1 and isinstance(val, (int, float)):
                c.number_format = '#,##0'; c.alignment = Alignment(horizontal="right")
    for i in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(i)].width = 22
    if rows:
        ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
        tbl = Table(displayName=tabell_navn, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tbl)

def bygg_excel(reg_rows, reg_innhold, kv_rows, kv_innhold):
    print(f"\nSparar Excel ({OUTPUT_EXCEL})...")
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    lag_ark(wb, "Region (arsvis)", reg_rows,
            ["Region", "Boligtype", "Tid"], reg_innhold, "TabellRegion", "SSB 06035")
    lag_ark(wb, "Nasjonalt (kvartal)", kv_rows,
            ["Boligtype", "Tid"], kv_innhold, "TabellKvartal", "SSB 07241")
    wb.save(OUTPUT_EXCEL)
    print(f"  Lagret: {OUTPUT_EXCEL}")

# ── KJØR ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    reg_rows, reg_innhold = hent_region()
    kv_rows,  kv_innhold  = hent_kvartal()
    spara_db(reg_rows, kv_rows)
    bygg_excel(reg_rows, reg_innhold, kv_rows, kv_innhold)
    print(f"\nFerdig!")
    print(f"  Database: {OUTPUT_DB}")
    print(f"  Excel:    {OUTPUT_EXCEL}")