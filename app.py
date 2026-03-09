"""
app.py – Streamlit-app for boligprisstatistikk
Norge (SSB) + Sverige (SCB) + Oslo færdigstilte boliger
"""

import sqlite3
import io
import requests
import streamlit as st
import pandas as pd
from itertools import product
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import plotly.graph_objects as go

DB_FIL          = "boligpriser.db"
URL_NO_REG      = "https://data.ssb.no/api/v0/en/table/06035"
URL_NO_KV       = "https://data.ssb.no/api/v0/en/table/07241"
URL_SE          = "https://api.scb.se/OV0104/v1/doris/sv/ssd/BO/BO0501/BO0501C/FastprisBRFRegionAr"
URL_OSLO_FF     = "https://data.ssb.no/api/v0/no/table/05889"
ONSKEDE_NO      = ["Oslo", "Bergen", "Trondheim", "Stavanger", "The whole country"]

st.set_page_config(page_title="Nordisk Boligprisstatistikk", page_icon="🏠", layout="wide")

# ── HJELPEFUNKSJONER ──────────────────────────────────────────────────────────
def fix(t):
    return t.replace("km²","m²").replace("km2","m²").replace("KM2","m²")

def hent_meta_no(url):
    r = requests.get(url, timeout=30); r.raise_for_status()
    return {v["code"]: {"values": v["values"], "labels": v["valueTexts"]}
            for v in r.json()["variables"]}

def hent_json_no(url, query):
    r = requests.post(url, json={"query": query, "response": {"format": "json-stat2"}}, timeout=30)
    r.raise_for_status(); return r.json()

def parse_no(data):
    dims = data["dimension"]; vals = data["value"]
    keys = list(dims.keys())
    items = [list(dims[k]["category"]["label"].items()) for k in keys]
    rows, idx = [], 0
    for combo in product(*items):
        v = vals[idx] if idx < len(vals) else None
        idx += 1
        if v is not None:
            row = {keys[i]: combo[i][1] for i in range(len(keys))}
            row["verdi"] = v
            rows.append(row)
    return rows

def hent_scb_data():
    meta = requests.get(URL_SE, timeout=30).json()
    regioner   = meta["variables"][0]["values"]
    tider      = meta["variables"][2]["values"][-8:]
    query = {
        "query": [
            {"code": "Region",       "selection": {"filter": "item", "values": regioner}},
            {"code": "ContentsCode", "selection": {"filter": "item", "values": ["BO0501R7"]}},
            {"code": "Tid",          "selection": {"filter": "item", "values": tider}},
        ],
        "response": {"format": "json-stat2"}
    }
    r = requests.post(URL_SE, json=query, timeout=30); r.raise_for_status()
    data = r.json()
    dims = data["dimension"]; vals = data["value"]
    reg_items = list(dims["Region"]["category"]["label"].items())
    tid_items = list(dims["Tid"]["category"]["label"].items())
    rows = []
    idx = 0
    for r_code, r_name in reg_items:
        for t_code, t_label in tid_items:
            v = vals[idx] if idx < len(vals) else None
            idx += 1
            if v is not None:
                rows.append({
                    "region": r_name, "bostadstyp": "Bostadsrätt",
                    "periode": t_label, "medelpris_tkr": v,
                    "land": "Sverige", "kilde": "SCB FastprisBRFRegionAr"
                })
    return rows

def hent_oslo_fullfort():
    """Hent fullførte boliger i Oslo fra SSB tabell 05889."""
    meta = hent_meta_no(URL_OSLO_FF)

    alle_bt = meta["Byggeareal"]["values"]
    alle_tid = meta["Tid"]["values"]

    query = [
        {"code": "Region",       "selection": {"filter": "item", "values": ["0301"]}},
        {"code": "Byggeareal",   "selection": {"filter": "item", "values": alle_bt}},
        {"code": "ContentsCode", "selection": {"filter": "item", "values": ["Fullforte"]}},
        {"code": "Tid",          "selection": {"filter": "item", "values": alle_tid}},
    ]
    data = hent_json_no(URL_OSLO_FF, query)
    rows = parse_no(data)

    # Summer over bygningstyper per kvartal
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    df["verdi"] = pd.to_numeric(df["verdi"], errors="coerce").fillna(0)
    df_agg = df.groupby("Tid")["verdi"].sum().reset_index()
    df_agg.columns = ["kvartal", "antall"]
    df_agg = df_agg.sort_values("kvartal").reset_index(drop=True)
    return df_agg

# ── OPPDATER DATABASE ─────────────────────────────────────────────────────────
def oppdater_db():
    meta_r = hent_meta_no(URL_NO_REG)
    reg_k = [k for k, l in zip(meta_r["Region"]["values"], meta_r["Region"]["labels"])
             if any(l.lower().startswith(o.lower()) or o.lower() in l.lower() for o in ONSKEDE_NO)]
    if not reg_k: reg_k = meta_r["Region"]["values"][:5]
    no_reg = parse_no(hent_json_no(URL_NO_REG, [
        {"code": "Region",       "selection": {"filter": "item", "values": reg_k}},
        {"code": "Boligtype",    "selection": {"filter": "item", "values": meta_r["Boligtype"]["values"][:3]}},
        {"code": "ContentsCode", "selection": {"filter": "item", "values": [meta_r["ContentsCode"]["values"][0]]}},
        {"code": "Tid",          "selection": {"filter": "item", "values": meta_r["Tid"]["values"][-8:]}},
    ]))

    meta_k = hent_meta_no(URL_NO_KV)
    no_kv = parse_no(hent_json_no(URL_NO_KV, [
        {"code": "Boligtype",    "selection": {"filter": "item", "values": meta_k["Boligtype"]["values"][:3]}},
        {"code": "ContentsCode", "selection": {"filter": "item", "values": [meta_k["ContentsCode"]["values"][0]]}},
        {"code": "Tid",          "selection": {"filter": "item", "values": meta_k["Tid"]["values"][-12:]}},
    ]))

    se_rows = hent_scb_data()
    oslo_df = hent_oslo_fullfort()

    conn = sqlite3.connect(DB_FIL); c = conn.cursor()

    c.execute("""CREATE TABLE IF NOT EXISTS region_arsvis (
        id INTEGER PRIMARY KEY AUTOINCREMENT, region TEXT, boligtype TEXT,
        periode TEXT, pris_m2 REAL, land TEXT, kilde TEXT, hentet TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS nasjonalt_kvartal (
        id INTEGER PRIMARY KEY AUTOINCREMENT, boligtype TEXT, periode TEXT,
        pris_m2 REAL, land TEXT DEFAULT 'Norge', kilde TEXT DEFAULT 'SSB 07241', hentet TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS sverige_brf (
        id INTEGER PRIMARY KEY AUTOINCREMENT, region TEXT, bostadstyp TEXT,
        periode TEXT, medelpris_tkr REAL, land TEXT, kilde TEXT, hentet TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS oslo_fullfort (
        id INTEGER PRIMARY KEY AUTOINCREMENT, kvartal TEXT, antall REAL, hentet TEXT)""")

    c.execute("DELETE FROM region_arsvis")
    c.execute("DELETE FROM nasjonalt_kvartal")
    c.execute("DELETE FROM sverige_brf")
    c.execute("DELETE FROM oslo_fullfort")

    hentet = date.today().isoformat()
    for row in no_reg:
        c.execute("INSERT INTO region_arsvis (region, boligtype, periode, pris_m2, land, kilde, hentet) VALUES (?,?,?,?,?,?,?)",
                  (row.get("Region"), row.get("Boligtype"), row.get("Tid"), row["verdi"], "Norge", "SSB 06035", hentet))
    for row in no_kv:
        c.execute("INSERT INTO nasjonalt_kvartal (boligtype, periode, pris_m2, hentet) VALUES (?,?,?,?)",
                  (row.get("Boligtype"), row.get("Tid"), row["verdi"], hentet))
    for row in se_rows:
        c.execute("INSERT INTO sverige_brf (region, bostadstyp, periode, medelpris_tkr, land, kilde, hentet) VALUES (?,?,?,?,?,?,?)",
                  (row["region"], row["bostadstyp"], row["periode"], row["medelpris_tkr"], row["land"], row["kilde"], hentet))
    if not oslo_df.empty:
        for _, row in oslo_df.iterrows():
            c.execute("INSERT INTO oslo_fullfort (kvartal, antall, hentet) VALUES (?,?,?)",
                      (row["kvartal"], row["antall"], hentet))

    conn.commit(); conn.close()
    return len(no_reg), len(no_kv), len(se_rows), len(oslo_df)

# ── HENT FRA DB ───────────────────────────────────────────────────────────────
@st.cache_data
def hent_no_region():
    conn = sqlite3.connect(DB_FIL)
    df = pd.read_sql("SELECT * FROM region_arsvis", conn)
    conn.close(); return df

@st.cache_data
def hent_no_kvartal():
    conn = sqlite3.connect(DB_FIL)
    df = pd.read_sql("SELECT * FROM nasjonalt_kvartal", conn)
    conn.close(); return df

@st.cache_data
def hent_se():
    conn = sqlite3.connect(DB_FIL)
    df = pd.read_sql("SELECT * FROM sverige_brf", conn)
    conn.close(); return df

@st.cache_data
def hent_oslo_ff():
    conn = sqlite3.connect(DB_FIL)
    try:
        df = pd.read_sql("SELECT * FROM oslo_fullfort ORDER BY kvartal", conn)
    except Exception:
        df = pd.DataFrame()
    conn.close(); return df

def lag_excel(df, arknavn):
    wb = Workbook(); ws = wb.active; ws.title = arknavn
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", start_color="1F4E79", fgColor="1F4E79")
    for col, h in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col, value=val)
    for i in range(1, len(df.columns)+1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    buffer = io.BytesIO(); wb.save(buffer); buffer.seek(0)
    return buffer

# ── LAYOUT ────────────────────────────────────────────────────────────────────
st.title("🏠 Nordisk Boligprisstatistikk")
col_tittel, col_knapp = st.columns([4, 1])
with col_knapp:
    if st.button("🔄 Hent ny data", use_container_width=True):
        with st.spinner("Henter data fra SSB og SCB..."):
            try:
                r, k, s, o = oppdater_db()
                st.cache_data.clear()
                st.success(f"Oppdatert! Norge: {r} reg.rader, {k} kv.rader · Sverige: {s} · Oslo fullført: {o} kvartaler")
            except Exception as e:
                st.error(f"Feil: {e}")
with col_tittel:
    st.caption("Norge: SSB (gratis) · Sverige: SCB (gratis)")

try:
    df_no_reg = hent_no_region()
    df_no_kv  = hent_no_kvartal()
    df_se     = hent_se()
    df_oslo   = hent_oslo_ff()
except Exception:
    st.warning("Ingen data funnet. Klikk 'Hent ny data' for å laste inn data.")
    st.stop()

tab1, tab2, tab3, tab4 = st.tabs([
    "🇳🇴 Norge – region (årsvis)",
    "🇳🇴 Norge – nasjonalt (kvartal)",
    "🇸🇪 Sverige – bostadsrätter (årsvis)",
    "🏗️ Oslo – færdigstilte boliger"
])

# ── TAB 1: NORGE REGION ───────────────────────────────────────────────────────
with tab1:
    col1, col2, col3 = st.columns(3)
    with col1:
        regioner = sorted(df_no_reg["region"].unique())
        valgte_reg = st.multiselect("Region", regioner, default=regioner)
    with col2:
        bt = sorted(df_no_reg["boligtype"].unique())
        valgte_bt = st.multiselect("Boligtype", bt, default=bt)
    with col3:
        perioder = sorted(df_no_reg["periode"].unique())
        valgt_p = st.selectbox("År", ["Alle"] + list(reversed(perioder)))
    filtrert = df_no_reg[df_no_reg["region"].isin(valgte_reg) & df_no_reg["boligtype"].isin(valgte_bt)]
    if valgt_p != "Alle": filtrert = filtrert[filtrert["periode"] == valgt_p]
    filtrert = filtrert[["region","boligtype","periode","pris_m2"]].sort_values(["periode","region"], ascending=[False,True]).rename(
        columns={"region":"Region","boligtype":"Boligtype","periode":"År","pris_m2":"Pris per m² (NOK)"})
    st.markdown(f"**{len(filtrert)} rader**")
    st.dataframe(filtrert.style.format({"Pris per m² (NOK)": "{:,.0f}"}), use_container_width=True, hide_index=True)
    st.download_button("⬇️ Last ned Excel", data=lag_excel(filtrert, "Norge region"),
                       file_name="norge_region.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    if valgt_p == "Alle" and valgte_reg:
        pivot = filtrert.pivot_table(index="År", columns="Region", values="Pris per m² (NOK)", aggfunc="mean")
        st.line_chart(pivot)

# ── TAB 2: NORGE KVARTAL ──────────────────────────────────────────────────────
with tab2:
    col1, col2 = st.columns(2)
    with col1:
        bt_kv = sorted(df_no_kv["boligtype"].unique())
        valgte_bt_kv = st.multiselect("Boligtype", bt_kv, default=bt_kv, key="bt_kv")
    with col2:
        antall = st.slider("Antall kvartaler", 4, 20, 12)
    perioder_kv = sorted(df_no_kv["periode"].unique())[-antall:]
    filtrert_kv = df_no_kv[df_no_kv["boligtype"].isin(valgte_bt_kv) & df_no_kv["periode"].isin(perioder_kv)
        ][["boligtype","periode","pris_m2"]].sort_values(["periode","boligtype"], ascending=[False,True]).rename(
        columns={"boligtype":"Boligtype","periode":"Kvartal","pris_m2":"Pris per m² (NOK)"})
    st.markdown(f"**{len(filtrert_kv)} rader**")
    st.dataframe(filtrert_kv.style.format({"Pris per m² (NOK)": "{:,.0f}"}), use_container_width=True, hide_index=True)
    st.download_button("⬇️ Last ned Excel", data=lag_excel(filtrert_kv, "Norge kvartal"),
                       file_name="norge_kvartal.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_kv")
    if valgte_bt_kv:
        pivot_kv = filtrert_kv.pivot_table(index="Kvartal", columns="Boligtype", values="Pris per m² (NOK)", aggfunc="mean")
        st.line_chart(pivot_kv)

# ── TAB 3: SVERIGE ────────────────────────────────────────────────────────────
with tab3:
    col1, col2 = st.columns(2)
    with col1:
        se_reg = sorted(df_se["region"].unique())
        valgte_se_reg = st.multiselect("Region (län)", se_reg, default=se_reg[:6])
    with col2:
        se_per = sorted(df_se["periode"].unique())
        valgt_se_p = st.selectbox("År", ["Alle"] + list(reversed(se_per)), key="se_p")
    filtrert_se = df_se[df_se["region"].isin(valgte_se_reg)]
    if valgt_se_p != "Alle": filtrert_se = filtrert_se[filtrert_se["periode"] == valgt_se_p]
    filtrert_se = filtrert_se[["region","periode","medelpris_tkr"]].sort_values(["periode","region"], ascending=[False,True]).rename(
        columns={"region":"Region","periode":"År","medelpris_tkr":"Medelpris (tkr)"})
    st.markdown(f"**{len(filtrert_se)} rader** · OBS: Medelpris i tusental kr, ikke kr/m²")
    st.dataframe(filtrert_se.style.format({"Medelpris (tkr)": "{:,.0f}"}), use_container_width=True, hide_index=True)
    st.download_button("⬇️ Ladda ned Excel", data=lag_excel(filtrert_se, "Sverige bostadsrätter"),
                       file_name="sverige_brf.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_se")
    if valgt_se_p == "Alle" and valgte_se_reg:
        pivot_se = filtrert_se.pivot_table(index="År", columns="Region", values="Medelpris (tkr)", aggfunc="mean")
        st.line_chart(pivot_se)

# ── TAB 4: OSLO FÆRDIGSTILTE BOLIGER ─────────────────────────────────────────
with tab4:
    st.subheader("Fullførte boliger i Oslo")
    st.caption("Kilde: SSB tabell 06265 · Alle bygningstyper summert · kv 1 2000 – siste kvartal")

    if df_oslo.empty:
        st.info("Ingen data. Klikk 'Hent ny data' øverst.")
    else:
        df_oslo = df_oslo.sort_values("kvartal").reset_index(drop=True)
        df_oslo["ma4"] = df_oslo["antall"].rolling(window=4).mean()
        snitt = df_oslo["antall"].mean()

        fig = go.Figure()

        # Stapeldiagram per kvartal
        fig.add_trace(go.Bar(
            x=df_oslo["kvartal"],
            y=df_oslo["antall"],
            name="Per kvartal",
            marker_color="#E8541A",
            opacity=0.85
        ))

        # 4-kvartalers glidende gjennomsnitt
        fig.add_trace(go.Scatter(
            x=df_oslo["kvartal"],
            y=df_oslo["ma4"],
            name="4-kv. glidende medelvärde",
            line=dict(color="#1A1A1A", width=2)
        ))

        # Gjennomsnittslinje (horisontal)
        fig.add_hline(
            y=snitt,
            line_color="#C0392B",
            line_width=1.5,
            annotation_text=f"Periodens genomsnitt: {snitt:.0f}",
            annotation_position="top left"
        )

        fig.update_layout(
            xaxis_title="Kvartal",
            yaxis_title="Antal boliger",
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
            plot_bgcolor="white",
            height=500,
            bargap=0.1
        )
        fig.update_xaxes(showgrid=False)
        fig.update_yaxes(showgrid=True, gridcolor="#EEEEEE")

        st.plotly_chart(fig, use_container_width=True)

        # Tabell + nedlasting
        vis_df = df_oslo[["kvartal","antall"]].rename(columns={"kvartal":"Kvartal","antall":"Antal fullförda"})
        st.download_button("⬇️ Last ned Excel", data=lag_excel(vis_df, "Oslo fullfort"),
                           file_name="oslo_fullfort.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_oslo")

st.divider()
st.caption("Prototype · Norge: SSB 06035 + 07241 + 06265 · Sverige: SCB FastprisBRFRegionAr · Klikk 'Hent ny data' for å oppdatere")
